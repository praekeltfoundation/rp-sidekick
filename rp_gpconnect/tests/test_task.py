import os
import tempfile
from unittest.mock import Mock, patch

import boto3
from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from moto import mock_s3
from openpyxl import Workbook

from rp_gpconnect.models import ContactImport, Flow
from rp_gpconnect.tasks import (
    import_or_update_contact,
    process_contact_import,
    pull_new_import_file,
)
from sidekick.models import Organization
from sidekick.tests.utils import assertCallMadeWith


def create_temp_xlsx_file(temp_file, msisdns):
    wb = Workbook()
    sheet = wb.create_sheet("GP Connect daily report", 0)
    sheet["A1"] = "msisdn"
    sheet["B1"] = "something_else"
    sheet["C1"] = "patients_tested_positive"
    for x in range(len(msisdns)):
        sheet.cell(row=(x + 2), column=1, value=msisdns[x])
        sheet.cell(row=(x + 2), column=2, value="stuuuuff")
        sheet.cell(row=(x + 2), column=3, value=(x % 2))
    wb.save(temp_file)
    return temp_file


@mock_s3
class PullNewImportFileTaskTests(TestCase):
    def setUp(self):
        self.org = Organization.objects.create(
            name="GP Connect", url="http://localhost:8002/", token="REPLACEME"
        )
        self.client = boto3.client("s3")
        self.client.create_bucket(Bucket="Test_Bucket")
        # Create imported dir
        self.imported_dir = os.path.join(tempfile.gettempdir(), "uploads/gpconnect")
        os.makedirs(self.imported_dir)

    def tearDown(self):
        s3 = boto3.resource("s3")
        bucket = s3.Bucket("Test_Bucket")
        for key in bucket.objects.all():
            key.delete()
        bucket.delete()
        # Cleanup filesystem
        for file in os.listdir(self.imported_dir):
            os.remove(os.path.join(self.imported_dir, file))
        os.rmdir(self.imported_dir)
        os.rmdir(os.path.join(tempfile.gettempdir(), "uploads"))

    @override_settings(
        MEDIA_ROOT=tempfile.gettempdir(), AWS_STORAGE_BUCKET_NAME="Test_Bucket"
    )
    @patch("rp_gpconnect.tasks.process_contact_import")
    def test_new_file_creates_contact_import_obj(self, mock_task):
        self.assertEqual(ContactImport.objects.count(), 0)
        with tempfile.NamedTemporaryFile() as temp_file:
            self.client.upload_file(
                Filename=temp_file.name,
                Bucket="Test_Bucket",
                Key="uploads/tempfile.xlsx",
            )

        pull_new_import_file(upload_dir="uploads/", org_name=self.org.name)
        self.assertEqual(ContactImport.objects.count(), 1)
        obj = ContactImport.objects.first()
        self.assertEqual(obj.file.name, "/tmp/uploads/gpconnect/tempfile.xlsx")
        mock_task.assert_called_with(obj.pk)

    @override_settings(
        MEDIA_ROOT=tempfile.gettempdir(), AWS_STORAGE_BUCKET_NAME="Test_Bucket"
    )
    @patch("rp_gpconnect.tasks.process_contact_import")
    def test_only_one_file_creates_contact_import_obj(self, mock_task):
        self.assertEqual(ContactImport.objects.count(), 0)
        with tempfile.NamedTemporaryFile() as temp_file:
            self.client.upload_file(
                Filename=temp_file.name,
                Bucket="Test_Bucket",
                Key="uploads/tempfile_1.xlsx",
            )
            self.client.upload_file(
                Filename=temp_file.name,
                Bucket="Test_Bucket",
                Key="uploads/tempfile_2.xlsx",
            )

        pull_new_import_file(upload_dir="uploads/", org_name=self.org.name)
        self.assertEqual(ContactImport.objects.count(), 1)
        obj = ContactImport.objects.first()
        existing_files = [
            "/tmp/uploads/gpconnect/tempfile_1.xlsx",
            "/tmp/uploads/gpconnect/tempfile_2.xlsx",
        ]
        self.assertIn(obj.file.name, existing_files)
        mock_task.assert_called()

    @override_settings(
        MEDIA_ROOT=tempfile.gettempdir(), AWS_STORAGE_BUCKET_NAME="Test_Bucket"
    )
    @patch("rp_gpconnect.tasks.process_contact_import")
    def test_already_processed_file_doesnt_create_contact_import_obj(self, mock_task):
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", dir=self.imported_dir
        ) as temp_file:
            filename = os.path.basename(temp_file.name)
            ContactImport.objects.create(file=temp_file.name, org=self.org)
            self.assertEqual(ContactImport.objects.count(), 1)
            self.client.upload_file(
                Filename=temp_file.name,
                Bucket="Test_Bucket",
                Key="uploads/{}".format(filename),
            )

        pull_new_import_file(upload_dir="uploads/", org_name=self.org.name)
        self.assertEqual(ContactImport.objects.count(), 1)
        mock_task.assert_not_called()

    @override_settings(
        MEDIA_ROOT=tempfile.gettempdir(), AWS_STORAGE_BUCKET_NAME="Test_Bucket"
    )
    @patch("rp_gpconnect.tasks.process_contact_import")
    def test_non_excel_files_and_prefix_are_ignored(self, mock_task):
        self.assertEqual(ContactImport.objects.count(), 0)
        with tempfile.NamedTemporaryFile() as temp_file:
            self.client.upload_file(
                Filename=temp_file.name, Bucket="Test_Bucket", Key="uploads/"
            )
            self.client.upload_file(
                Filename=temp_file.name,
                Bucket="Test_Bucket",
                Key="uploads/non_excel.txt",
            )

        pull_new_import_file(upload_dir="uploads/", org_name=self.org.name)
        self.assertEqual(ContactImport.objects.count(), 0)
        mock_task.assert_not_called()


class ProcessContactImportTaskTests(TestCase):
    def setUp(self):
        self.org = Organization.objects.create(
            name="Test Organization", url="http://localhost:8002/", token="REPLACEME"
        )
        self.user = User.objects.create_user(
            "username", "testuser@example.com", "password"
        )

    @override_settings(MEDIA_ROOT=tempfile.gettempdir())
    @patch("rp_gpconnect.tasks.log")
    @patch("rp_gpconnect.tasks.import_or_update_contact.delay")
    def test_contact_import_calls_contact_task_for_each_positive_row(
        self, mock_contact_update_task, mock_logger
    ):
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx")
        contacts_file = create_temp_xlsx_file(
            temp_file, ["+27000000001", "+27000000002", "+27000000003", "+27000000004"]
        ).name
        import_obj = ContactImport.objects.create(
            file=contacts_file, org=self.org, created_by=self.user
        )

        process_contact_import(import_obj.pk)

        mock_logger.info.assert_called_with(
            "Importing contacts for file: %s" % contacts_file
        )

        self.assertEqual(mock_contact_update_task.call_count, 2)
        mock_contact_update_task.assert_any_call(
            {
                "msisdn": "+27000000002",
                "something_else": "stuuuuff",
                "patients_tested_positive": 1,
            },
            self.org.pk,
        )
        mock_contact_update_task.assert_any_call(
            {
                "msisdn": "+27000000004",
                "something_else": "stuuuuff",
                "patients_tested_positive": 1,
            },
            self.org.pk,
        )


class ImportOrUpdateContactTaskTests(TestCase):
    def setUp(self):
        self.org = Organization.objects.create(
            name="Test Organization", url="http://localhost:8002/", token="REPLACEME"
        )
        self.create_flow = Flow.objects.create(
            type="new_contact", rapidpro_flow="new_contact_flow", org=self.org
        )
        self.update_flow = Flow.objects.create(
            type="contact_update", rapidpro_flow="contact_update_flow", org=self.org
        )
        self.user = User.objects.create_user(
            "username", "testuser@example.com", "password"
        )

    @patch("temba_client.v2.TembaClient.create_flow_start")
    @patch("temba_client.v2.TembaClient.update_contact")
    @patch("temba_client.v2.TembaClient.get_contacts")
    @patch("rp_gpconnect.tasks.get_whatsapp_contact_id")
    def test_contact_task_if_contact_exists_without_wa_id(
        self,
        mock_get_whatsapp_contact_id,
        mock_get_rp_contact,
        mock_update_rp_contact,
        mock_create_flow_start,
    ):
        mock_get_whatsapp_contact_id.return_value = None

        # Create a mock contact that is already up to date
        mock_contact_object = Mock()
        mock_contact_object.uuid = "123456"
        mock_contact_object.urns = ["tel:+27000000001"]
        mock_contact_object.fields = {
            "something_else": "stuuuuff",
            "has_whatsapp": False,
        }
        mock_get_rp_contact.return_value.first.return_value = mock_contact_object

        import_or_update_contact(
            {"msisdn": "+27000000001", "something_else": "stuuuuff"}, self.org.pk
        )

        mock_get_whatsapp_contact_id.assert_called_with(self.org, "+27000000001")
        self.assertEqual(mock_get_rp_contact.call_count, 1)
        mock_update_rp_contact.assert_not_called()
        mock_create_flow_start.assert_not_called()

    @patch("temba_client.v2.TembaClient.create_flow_start")
    @patch("temba_client.v2.TembaClient.create_contact")
    @patch("temba_client.v2.TembaClient.update_contact")
    @patch("temba_client.v2.TembaClient.get_contacts")
    @patch("rp_gpconnect.tasks.get_whatsapp_contact_id")
    def test_contact_task_if_contact_exists_with_wa_id(
        self,
        mock_get_whatsapp_contact_id,
        mock_get_rp_contact,
        mock_update_rp_contact,
        mock_create_rp_contact,
        mock_create_flow_start,
    ):
        mock_get_whatsapp_contact_id.return_value = "27000000001"

        mock_contact_object = Mock()
        mock_contact_object.uuid = "123456"
        mock_contact_object.urns = ["tel:+27000000001"]
        mock_get_rp_contact.return_value.first.return_value = mock_contact_object

        import_or_update_contact(
            {"msisdn": "+27000000001", "something_else": "stuuuuff"}, self.org.pk
        )
        mock_get_whatsapp_contact_id.assert_called_with(self.org, "+27000000001")

        self.assertEqual(mock_get_rp_contact.call_count, 1)
        mock_update_rp_contact.assert_called_with(
            contact="123456", urns=["tel:+27000000001", "whatsapp:27000000001"]
        )
        assertCallMadeWith(
            mock_create_flow_start.call_args,
            flow=self.update_flow.rapidpro_flow,
            urns=["tel:+27000000001", "whatsapp:27000000001"],
            restart_participants=True,
            extra={"something_else": "stuuuuff", "has_whatsapp": True},
        )

        mock_create_rp_contact.assert_not_called()

    @patch("temba_client.v2.TembaClient.create_flow_start", autospec=True)
    @patch("temba_client.v2.TembaClient.create_contact", autospec=True)
    @patch("temba_client.v2.TembaClient.update_contact", autospec=True)
    @patch("temba_client.v2.TembaClient.get_contacts", autospec=True)
    @patch("rp_gpconnect.tasks.get_whatsapp_contact_id", autospec=True)
    def test_contact_task_if_new_contact_without_wa_id(
        self,
        mock_get_whatsapp_contact_id,
        mock_get_rp_contact,
        mock_update_rp_contact,
        mock_create_rp_contact,
        mock_create_flow_start,
    ):
        mock_get_whatsapp_contact_id.return_value = None

        mock_get_rp_contact.return_value.first.side_effect = [None, None]

        import_or_update_contact(
            {"msisdn": "+27000000001", "something_else": "stuuuuff"}, self.org.pk
        )

        self.assertEqual(mock_get_rp_contact.call_count, 1)
        mock_update_rp_contact.assert_not_called()

        assertCallMadeWith(mock_create_rp_contact.call_args, urns=["tel:+27000000001"])

        assertCallMadeWith(
            mock_create_flow_start.call_args,
            flow=self.create_flow.rapidpro_flow,
            urns=["tel:+27000000001"],
            restart_participants=True,
            extra={"something_else": "stuuuuff", "has_whatsapp": False},
        )

    @patch("temba_client.v2.TembaClient.create_flow_start", autospec=True)
    @patch("temba_client.v2.TembaClient.create_contact", autospec=True)
    @patch("temba_client.v2.TembaClient.update_contact", autospec=True)
    @patch("temba_client.v2.TembaClient.get_contacts", autospec=True)
    @patch("rp_gpconnect.tasks.get_whatsapp_contact_id", autospec=True)
    def test_contact_task_if_new_contact(
        self,
        mock_get_whatsapp_contact_id,
        mock_get_rp_contact,
        mock_update_rp_contact,
        mock_create_rp_contact,
        mock_create_flow_start,
    ):
        mock_get_whatsapp_contact_id.return_value = "27000000001"

        mock_get_rp_contact.return_value.first.side_effect = [None, None]

        import_or_update_contact(
            {"msisdn": "+27000000001", "something_else": "stuuuuff"}, self.org.pk
        )

        self.assertEqual(mock_get_rp_contact.call_count, 2)
        call_1_args, call_2_args = mock_get_rp_contact.call_args_list
        assertCallMadeWith(call_1_args, urn="tel:+27000000001")
        assertCallMadeWith(call_2_args, urn="whatsapp:27000000001")

        mock_update_rp_contact.assert_not_called()

        assertCallMadeWith(
            mock_create_rp_contact.call_args,
            urns=["tel:+27000000001", "whatsapp:27000000001"],
        )

        assertCallMadeWith(
            mock_create_flow_start.call_args,
            flow=self.create_flow.rapidpro_flow,
            urns=["tel:+27000000001", "whatsapp:27000000001"],
            restart_participants=True,
            extra={"something_else": "stuuuuff", "has_whatsapp": True},
        )
